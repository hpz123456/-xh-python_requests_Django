# coding=utf-8
#Excel的读写

import openpyxl
from realize_package.model.Excel_message import Excel
import os
# curPath = os.path.abspath(os.path.dirname(__file__))
# rootPath = curPath[:curPath.find("InterfaceAutomationTest\\")+len("InterfaceAutomationTest\\")]  # 获取myProject，也就是项目的根路径
import sys
rootPath = ""
for i in (sys.path[0].split("\\")):
    rootPath = rootPath + i + "/"
    if i == "InterfaceAutomationTest":
        break

class operation_Excel():
    Files_Path = "realize_package/File_data/"
    # 打开文件
    def open_Excel(self,file_path):
        # print("Excel" + file_path)
        Excle_File = openpyxl.load_workbook(file_path)
        return Excle_File

    # 读Excel的某一个单元格
    def read_one_cell(self,sheet,row,col):
        return sheet.cell(row,col)

    #获取传入的sheet页的所有信息
    def all_Information(self,sheet):
        row_col = {
            "max_row":sheet.max_row,
            "max_columns":sheet.max_column
        }
        return row_col

    # 通过名称获取sheet
    def get_sheet(self,Excle_File,sheet_name):
        # read_excel(file_path, sheet_name=sheetName)
        return Excle_File[sheet_name]

    #获取首行的名字及（行，列）
    def get_fiest_line(self,sheet,max_col):
        Excel_list = []
        i = 1
        while i <= max_col:
            Excel_line = Excel()
            Excel_line.first_line_name = sheet.cell(1,i).value
            Excel_line.col = i
            Excel_line.row = 1
            Excel_list.append(Excel_line)
            i = i + 1

        return Excel_list

    #获取所有列的名字及（行，列）
    def get_All_row(self,sheet,max_row,col):
        Excel_list = []
        i = 1
        while i <= max_row:
            Excel_line = Excel()
            Excel_line.first_line_name = sheet.cell(i, col).value
            Excel_line.col = col
            Excel_line.row = i
            Excel_list.append(Excel_line)
            i = i + 1

        return Excel_list

    #写入Excel文件中
    def write_Excel(self,sheet,row,col,content):
        # print(sheet)
        # print(row)
        # print(col)
        # print(type(content))
        sheet.cell(row = row, column = col, value = str(content))

    #保存文件
    def save_File(self,Excle_File,file_path):
        Excle_File.save(file_path)

    #关闭文件
    def clos_Excel(self,Excle_File):
        Excle_File.exit()